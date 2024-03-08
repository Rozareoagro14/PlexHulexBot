import os
import requests
from openpyxl import load_workbook
from config import PLEX_SERVER_URL, X_PLEX_TOKEN, LIBRARY_SECTION_ID, MOVIES_DIRECTORY

def create_movie_placeholder(movie_title):
    """Создает файл-заглушку для фильма."""
    movie_file_name = f"{movie_title}.mp4"
    movie_file_path = os.path.join(MOVIES_DIRECTORY, movie_file_name)
    with open(movie_file_path, 'w') as f:
        f.write("file_id: UOIDuoudoDUO")

def scan_plex_library():
    """Инициирует сканирование библиотеки Plex."""
    headers = {'X-Plex-Token': X_PLEX_TOKEN}
    url = f"{PLEX_SERVER_URL}/library/sections/{LIBRARY_SECTION_ID}/refresh"
    response = requests.get(url, headers=headers)
    if response.status_code == 200:
        print("Библиотека обновляется...")
    else:
        print(f"Ошибка при обновлении библиотеки. Код ответа: {response.status_code}")

def process_movies_list():
    """Читает названия фильмов из Excel файла и создает для каждого заглушку."""
    movies_list_file = os.path.join(MOVIES_DIRECTORY, 'BD_film.xlsx')
    wb = load_workbook(filename=movies_list_file)
    sheet = wb.active  # Предполагаем, что данные находятся на первом листе
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Пропускаем заголовок
        movie_title = row[0]  # Предполагаем, что названия фильмов находятся в первой колонке
        if movie_title:  # Проверяем, что ячейка не пуста
            create_movie_placeholder(movie_title)
    scan_plex_library()

if __name__ == "__main__":
    process_movies_list()
